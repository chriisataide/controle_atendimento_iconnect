"""
Views de Compliance — Trilha de Auditoria e LGPD.

Features:
    - AuditTrailView: Painel interativo de trilha de auditoria (BACEN-compliant)
    - LGPDPanelView: Painel de gestão LGPD com consentimentos e solicitações
    - API endpoints para filtros/export
"""
import json
from datetime import timedelta

from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views import View
from django.views.generic import ListView
from django.db.models import Q, Count
from django.contrib.auth.models import User

from dashboard.models.audit import AuditEvent, SecurityAlert
from dashboard.models.lgpd import LGPDConsent, LGPDDataRequest, LGPDAccessLog


# ========== TRILHA DE AUDITORIA ==========

@method_decorator([login_required, staff_member_required], name='dispatch')
class AuditTrailView(ListView):
    """Trilha de auditoria com filtros avançados — BACEN/LGPD compliant."""
    model = AuditEvent
    template_name = 'dashboard/compliance/audit_trail.html'
    context_object_name = 'events'
    paginate_by = 50

    def get_queryset(self):
        qs = AuditEvent.objects.select_related('user').all()

        # Filtros
        event_type = self.request.GET.get('type')
        severity = self.request.GET.get('severity')
        user_id = self.request.GET.get('user')
        search = self.request.GET.get('q')
        days = self.request.GET.get('days')
        suspicious = self.request.GET.get('suspicious')

        if event_type:
            qs = qs.filter(event_type=event_type)
        if severity:
            qs = qs.filter(severity=severity)
        if user_id:
            qs = qs.filter(user_id=user_id)
        if search:
            qs = qs.filter(
                Q(action__icontains=search) |
                Q(description__icontains=search) |
                Q(ip_address__icontains=search) |
                Q(user__username__icontains=search)
            )
        if days:
            try:
                since = timezone.now() - timedelta(days=int(days))
                qs = qs.filter(timestamp__gte=since)
            except (ValueError, TypeError):
                pass
        if suspicious == '1':
            qs = qs.filter(is_suspicious=True)

        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        now = timezone.now()
        last_24h = now - timedelta(hours=24)
        last_7d = now - timedelta(days=7)
        last_30d = now - timedelta(days=30)

        total_all = AuditEvent.objects.count()
        total_30d = AuditEvent.objects.filter(timestamp__gte=last_30d).count()

        ctx['stats'] = {
            'total_24h': AuditEvent.objects.filter(timestamp__gte=last_24h).count(),
            'total_7d': AuditEvent.objects.filter(timestamp__gte=last_7d).count(),
            'total_30d': total_30d,
            'total_all': total_all,
            'suspicious': AuditEvent.objects.filter(is_suspicious=True, is_resolved=False).count(),
            'security_alerts': SecurityAlert.objects.exclude(status='resolved').count(),
        }
        ctx['event_types'] = AuditEvent.EVENT_TYPES
        ctx['severity_levels'] = AuditEvent.SEVERITY_LEVELS
        ctx['users'] = User.objects.filter(is_active=True).order_by('username')

        # Dados para chart de eventos por tipo (últimos 7d)
        type_counts = (
            AuditEvent.objects.filter(timestamp__gte=last_7d)
            .values('event_type')
            .annotate(count=Count('id'))
            .order_by('-count')
        )
        ctx['chart_labels'] = json.dumps([t['event_type'] for t in type_counts])
        ctx['chart_data'] = json.dumps([t['count'] for t in type_counts])

        # Tendência diária (últimos 7 dias)
        from django.db.models.functions import TruncDate
        trend_qs = (
            AuditEvent.objects.filter(timestamp__gte=last_7d)
            .annotate(dia=TruncDate('timestamp'))
            .values('dia')
            .annotate(count=Count('id'))
            .order_by('dia')
        )
        trend_map = {item['dia']: item['count'] for item in trend_qs}
        trend_labels = []
        trend_data = []
        for i in range(7):
            day = (now - timedelta(days=6 - i)).date()
            trend_labels.append(day.strftime('%d/%m'))
            trend_data.append(trend_map.get(day, 0))
        ctx['trend_labels'] = json.dumps(trend_labels)
        ctx['trend_data'] = json.dumps(trend_data)

        # Contagem por severidade (para doughnut)
        sev_counts = (
            AuditEvent.objects.filter(timestamp__gte=last_7d)
            .values('severity')
            .annotate(count=Count('id'))
            .order_by('-count')
        )
        sev_map = {s['severity']: s['count'] for s in sev_counts}
        ctx['sev_labels'] = json.dumps(['Baixa', 'Média', 'Alta', 'Crítica'])
        ctx['sev_data'] = json.dumps([
            sev_map.get('low', 0),
            sev_map.get('medium', 0),
            sev_map.get('high', 0),
            sev_map.get('critical', 0),
        ])

        # Filtros ativos (para exibir tags)
        ctx['active_filters'] = {
            k: v for k, v in self.request.GET.items() if v and k != 'page'
        }
        return ctx


@login_required
@staff_member_required
def audit_export_csv(request):
    """Exporta eventos de auditoria em Excel com layout padronizado."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from dashboard.views.features import _apply_excel_style, _auto_adjust_columns

    wb = Workbook()
    ws = wb.active
    ws.title = "Auditoria"

    headers = [
        "Data/Hora", "Tipo", "Severidade", "Usuário", "Ação",
        "Descrição", "IP", "Suspeito", "Resolvido",
    ]

    data_start_row, thin_border = _apply_excel_style(
        ws, headers, header_fill_color="475569",
        title_row="Relatório de Auditoria — iConnect",
    )

    zebra_fill = PatternFill(start_color="F2F6FC", end_color="F2F6FC", fill_type="solid")
    data_alignment = Alignment(vertical="center")

    qs = AuditEvent.objects.select_related('user').order_by('-timestamp')[:5000]

    row_num = data_start_row + 1
    for e in qs.iterator():
        values = [
            e.timestamp.strftime('%d/%m/%Y %H:%M:%S'),
            e.get_event_type_display(),
            e.get_severity_display(),
            e.user.username if e.user else '—',
            e.action,
            e.description[:200],
            e.ip_address,
            'Sim' if e.is_suspicious else 'Não',
            'Sim' if e.is_resolved else 'Não',
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = data_alignment
            if row_num % 2 == 0:
                cell.fill = zebra_fill
        row_num += 1

    _auto_adjust_columns(ws)
    ws.freeze_panes = ws.cell(row=data_start_row + 1, column=1)
    ws.auto_filter.ref = f"A{data_start_row}:{chr(64 + len(headers))}{row_num - 1}"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"auditoria_{timezone.now():%Y%m%d_%H%M}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ========== PAINEL LGPD ==========

@method_decorator([login_required, staff_member_required], name='dispatch')
class LGPDPanelView(View):
    """Painel de gestão LGPD — consentimentos, solicitações e logs."""
    template_name = 'dashboard/compliance/lgpd_panel.html'

    def get(self, request):
        now = timezone.now()
        last_30d = now - timedelta(days=30)

        # Stats
        total_consents = LGPDConsent.objects.filter(granted=True, revoked_at__isnull=True).count()
        pending_requests = LGPDDataRequest.objects.filter(status='pending').count()
        overdue_requests = LGPDDataRequest.objects.filter(
            status__in=['pending', 'in_progress'],
            deadline__lt=now,
        ).count()
        access_logs_30d = LGPDAccessLog.objects.filter(timestamp__gte=last_30d).count()

        # Solicitações recentes
        recent_requests = (
            LGPDDataRequest.objects
            .select_related('user', 'processed_by')
            .order_by('-created_at')[:20]
        )

        # Distribuição de consentimentos por finalidade
        consent_stats = (
            LGPDConsent.objects
            .filter(granted=True, revoked_at__isnull=True)
            .values('purpose')
            .annotate(count=Count('id'))
            .order_by('-count')
        )

        # Logs de acesso recentes
        recent_logs = (
            LGPDAccessLog.objects
            .select_related('user')
            .order_by('-timestamp')[:15]
        )

        context = {
            'stats': {
                'total_consents': total_consents,
                'pending_requests': pending_requests,
                'overdue_requests': overdue_requests,
                'access_logs_30d': access_logs_30d,
            },
            'recent_requests': recent_requests,
            'consent_stats': consent_stats,
            'recent_logs': recent_logs,
            'consent_purposes': LGPDConsent.PURPOSE_CHOICES,
            'request_types': LGPDDataRequest.REQUEST_TYPES,
        }
        return render(request, self.template_name, context)


@login_required
@staff_member_required
def lgpd_process_request(request, pk):
    """Processa solicitação LGPD (aprovar/rejeitar)."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        lgpd_request = LGPDDataRequest.objects.get(pk=pk)
    except LGPDDataRequest.DoesNotExist:
        return JsonResponse({'error': 'Solicitação não encontrada'}, status=404)

    action = request.POST.get('action')
    response_text = request.POST.get('response', '')

    if action == 'approve':
        lgpd_request.complete(processed_by=request.user, response=response_text)
        return JsonResponse({'status': 'approved', 'message': 'Solicitação aprovada com sucesso.'})
    elif action == 'reject':
        lgpd_request.reject(processed_by=request.user, response=response_text)
        return JsonResponse({'status': 'rejected', 'message': 'Solicitação rejeitada.'})
    else:
        return JsonResponse({'error': 'Ação inválida'}, status=400)
