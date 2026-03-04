"""
Management command para importação massiva de Pontos de Venda
a partir da planilha 'Pontos de vendas.xlsx'.

Importa TODOS os registros (qualquer status).
Para linhas sem CNPJ, gera CNPJ fictício baseado no UNIORG.
Para CNPJs duplicados, adiciona sufixo incremental.

Uso:
    python manage.py importar_pdv            # Importa apenas se não existirem PDVs
    python manage.py importar_pdv --force    # Remove PDVs existentes e reimporta tudo
"""

import os
import re

import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand
from django.db import transaction

from dashboard.models import Cliente, PontoDeVenda

# --- Configuração ---
NOME_CLIENTE = "Santander"
EMAIL_CLIENTE = "contato@santander.com.br"

# Placeholders para campos obrigatórios ausentes na planilha
PLACEHOLDER_EMAIL = "pdv@santander.com.br"
PLACEHOLDER_CELULAR = "(00) 00000-0000"
PLACEHOLDER_RESPONSAVEL_NOME = "A definir"
PLACEHOLDER_RESPONSAVEL_CPF = "000.000.000-00"
PLACEHOLDER_RESPONSAVEL_CARGO = "A definir"
PLACEHOLDER_RESPONSAVEL_TELEFONE = "(00) 00000-0000"
PLACEHOLDER_RESPONSAVEL_EMAIL = "responsavel@santander.com.br"


def formatar_cnpj(cnpj_raw):
    """Formata CNPJ para o padrão XX.XXX.XXX/XXXX-XX."""
    if not cnpj_raw:
        return ""
    digits = re.sub(r"\D", "", str(cnpj_raw))
    if not digits:
        return ""
    digits = digits.zfill(14)
    digits = digits[-14:]
    return "{}.{}.{}/{}-{}".format(
        digits[:2], digits[2:5], digits[5:8], digits[8:12], digits[12:14]
    )


def gerar_cnpj_ficticio(uniorg, idx):
    """Gera um CNPJ fictício baseado no UNIORG."""
    digits = re.sub(r"\D", "", str(uniorg))
    if not digits:
        digits = str(idx)
    digits = digits.zfill(8)[:8]
    filial = str(idx).zfill(4)[:4]
    check = "99"
    cnpj = "{}.{}.{}/{}-{}".format(
        digits[:2], digits[2:5], digits[5:8], filial, check
    )
    return cnpj


def formatar_cep(cep_raw):
    """Formata CEP para XXXXX-XXX."""
    if not cep_raw:
        return "00000-000"
    digits = re.sub(r"\D", "", str(cep_raw))
    digits = digits.zfill(8)
    return "{}-{}".format(digits[:5], digits[5:8])


def separar_logradouro_numero(endereco):
    """Separa número do logradouro."""
    if not endereco:
        return ("S/N", "S/N")
    endereco = str(endereco).strip()
    match = re.match(r"^(.+?)\s+(\d+[\w\-]*)$", endereco)
    if match:
        return (match.group(1).strip(), match.group(2).strip())
    return (endereco, "S/N")


def formatar_telefone(fone, ddd):
    """Formata telefone com DDD."""
    if not fone:
        return ""
    fone_str = str(fone).strip()
    ddd_str = str(int(ddd)) if ddd else ""
    if ddd_str:
        return "({}) {}".format(ddd_str, fone_str)
    return fone_str


class Command(BaseCommand):
    help = "Importa Pontos de Venda a partir da planilha Excel em data/Pontos de vendas.xlsx"

    def add_arguments(self, parser):
        parser.add_argument(
            "--force",
            action="store_true",
            help="Remove todos os PDVs existentes do cliente e reimporta tudo.",
        )
        parser.add_argument(
            "--planilha",
            type=str,
            default=None,
            help="Caminho para a planilha Excel (padrão: data/Pontos de vendas.xlsx).",
        )

    def handle(self, *args, **options):
        force = options["force"]
        planilha = options["planilha"] or os.path.join(
            settings.BASE_DIR, "data", "Pontos de vendas.xlsx"
        )

        self.stdout.write("=" * 60)
        self.stdout.write("  IMPORTAÇÃO DE PONTOS DE VENDA")
        self.stdout.write("=" * 60)

        # Verificar se a planilha existe
        if not os.path.exists(planilha):
            self.stdout.write(
                self.style.WARNING(
                    f"\n  Planilha não encontrada: {planilha}"
                )
            )
            self.stdout.write(
                self.style.WARNING("  Importação de PDVs ignorada.")
            )
            return

        # Criar ou buscar Cliente
        cliente, created = Cliente.objects.get_or_create(
            nome=NOME_CLIENTE,
            defaults={
                "email": EMAIL_CLIENTE,
                "segmento": "Financeiro",
                "empresa": "Santander Brasil",
                "ativo": True,
            },
        )
        if created:
            self.stdout.write(
                self.style.SUCCESS(
                    f'\n  Cliente "{NOME_CLIENTE}" criado (ID={cliente.id})'
                )
            )
        else:
            self.stdout.write(f'\n  Cliente "{NOME_CLIENTE}" já existe (ID={cliente.id})')

        # Verificar se já existem PDVs (idempotência)
        existing_count = PontoDeVenda.objects.filter(cliente=cliente).count()

        if existing_count > 0 and not force:
            self.stdout.write(
                self.style.SUCCESS(
                    f"\n  ✅ {existing_count} PDVs já existem para {NOME_CLIENTE}. "
                    f"Importação ignorada (use --force para reimportar)."
                )
            )
            return

        if existing_count > 0 and force:
            self.stdout.write(
                f"  Removendo {existing_count} PDVs existentes (--force)..."
            )
            PontoDeVenda.objects.filter(cliente=cliente).delete()
            self.stdout.write("  Removidos com sucesso.")

        # Abrir planilha
        self.stdout.write(f"\n  Abrindo planilha: {planilha}")
        wb = openpyxl.load_workbook(planilha, read_only=True)
        ws = wb.active
        total_linhas = ws.max_row - 1
        self.stdout.write(f"  Total de linhas: {total_linhas}")

        # Ler e preparar dados
        cnpjs_usados = set(PontoDeVenda.objects.values_list("cnpj", flat=True))
        pdvs_para_criar = []
        cnpj_gerado_count = 0
        cnpj_dedup_count = 0
        erros = []

        self.stdout.write("\n  Processando linhas...")
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            uniorg = str(row[0] or "").strip()
            tipo = str(row[1] or "").strip()
            nome = str(row[2] or "").strip()
            endereco = str(row[3] or "").strip()
            cep_raw = str(row[4] or "").strip()
            bairro = str(row[5] or "").strip()
            municipio = str(row[6] or "").strip()
            uf = str(row[7] or "").strip()
            cnpj_raw = str(row[8] or "").strip()
            fone = row[9]
            ddd = row[10]
            status = str(row[11] or "").strip()
            str(row[12] or "").strip()  # coluna ignorada
            rede = str(row[13] or "").strip()
            str(row[14] or "").strip()  # coluna ignorada
            str(row[15] or "").strip()  # coluna ignorada

            # Formatar CNPJ ou gerar fictício
            cnpj = formatar_cnpj(cnpj_raw)
            if not cnpj or cnpj == "00.000.000/0000-00":
                cnpj = gerar_cnpj_ficticio(uniorg, idx)
                cnpj_gerado_count += 1

            # Resolver duplicatas adicionando sufixo
            original_cnpj = cnpj
            suffix = 1
            while cnpj in cnpjs_usados:
                base = original_cnpj[:-2]
                cnpj = "{}{:02d}".format(base, suffix)
                suffix += 1
                cnpj_dedup_count += 1

            cnpjs_usados.add(cnpj)

            # Separar logradouro e número
            logradouro, numero = separar_logradouro_numero(endereco)

            # Montar nome fantasia
            nome_fantasia = nome if nome else "{} {}".format(tipo, uniorg).strip()
            razao_social = (
                "{} ({} {})".format(nome, tipo, uniorg).strip()
                if nome
                else nome_fantasia
            )

            # Telefone formatado
            telefone = formatar_telefone(fone, ddd)

            # Complemento com tipo e rede
            complemento_parts = [tipo]
            if rede:
                complemento_parts.append(rede)
            comp = " | ".join(complemento_parts)

            try:
                pdv = PontoDeVenda(
                    cliente=cliente,
                    razao_social=razao_social[:150],
                    nome_fantasia=nome_fantasia[:150],
                    cnpj=cnpj,
                    inscricao_estadual=uniorg[:30],
                    inscricao_municipal=status[:30],
                    cep=formatar_cep(cep_raw),
                    logradouro=logradouro[:120],
                    numero=numero[:10],
                    complemento=comp[:50],
                    bairro=bairro[:60] if bairro else "N/I",
                    cidade=municipio[:60] if municipio else "N/I",
                    estado=uf[:2] if uf else "XX",
                    pais="Brasil",
                    celular=telefone if telefone else PLACEHOLDER_CELULAR,
                    email_principal=PLACEHOLDER_EMAIL,
                    email_financeiro="",
                    website="",
                    responsavel_nome=PLACEHOLDER_RESPONSAVEL_NOME,
                    responsavel_cpf=PLACEHOLDER_RESPONSAVEL_CPF,
                    responsavel_cargo=PLACEHOLDER_RESPONSAVEL_CARGO,
                    responsavel_telefone=PLACEHOLDER_RESPONSAVEL_TELEFONE,
                    responsavel_email=PLACEHOLDER_RESPONSAVEL_EMAIL,
                )
                pdvs_para_criar.append(pdv)
            except Exception as e:
                erros.append("Linha {}: {}".format(idx, e))

        self.stdout.write(f"  Pronto para importar: {len(pdvs_para_criar)}")
        self.stdout.write(f"  CNPJs gerados (fictícios): {cnpj_gerado_count}")
        self.stdout.write(f"  CNPJs deduplicados: {cnpj_dedup_count}")
        if erros:
            self.stdout.write(self.style.WARNING(f"  Erros: {len(erros)}"))
            for e in erros[:10]:
                self.stdout.write(f"    ERRO: {e}")

        if not pdvs_para_criar:
            self.stdout.write(
                self.style.WARNING("\n  Nenhum ponto de venda para importar.")
            )
            wb.close()
            return

        # Criptografar campos PII antes do bulk_create
        self.stdout.write("\n  Criptografando campos PII...")
        try:
            from dashboard.utils.crypto import encrypt_value

            for pdv in pdvs_para_criar:
                if pdv.responsavel_cpf and not pdv.responsavel_cpf.startswith(
                    "enc::"
                ):
                    pdv.responsavel_cpf = encrypt_value(pdv.responsavel_cpf)
                if pdv.celular and not pdv.celular.startswith("enc::"):
                    pdv.celular = encrypt_value(pdv.celular)
                if pdv.responsavel_telefone and not pdv.responsavel_telefone.startswith(
                    "enc::"
                ):
                    pdv.responsavel_telefone = encrypt_value(
                        pdv.responsavel_telefone
                    )
        except ImportError:
            self.stdout.write(
                self.style.WARNING(
                    "  Módulo crypto não disponível, salvando sem criptografia"
                )
            )

        # Inserir em lote
        self.stdout.write(f"\n  Inserindo {len(pdvs_para_criar)} pontos de venda...")
        batch_size = 500
        total_criados = 0

        with transaction.atomic():
            for i in range(0, len(pdvs_para_criar), batch_size):
                batch = pdvs_para_criar[i : i + batch_size]
                PontoDeVenda.objects.bulk_create(batch, ignore_conflicts=True)
                total_criados += len(batch)
                self.stdout.write(
                    f"  ... {total_criados}/{len(pdvs_para_criar)} inseridos"
                )

        wb.close()

        # Resumo
        total_final = PontoDeVenda.objects.count()
        self.stdout.write("\n" + "=" * 60)
        self.stdout.write("  RESUMO DA IMPORTAÇÃO")
        self.stdout.write("=" * 60)
        self.stdout.write(
            self.style.SUCCESS(
                f"  ✅ Pontos de venda criados: {len(pdvs_para_criar)}"
            )
        )
        self.stdout.write(f"  Total de PDVs no banco: {total_final}")
        self.stdout.write(f"  CNPJs fictícios gerados: {cnpj_gerado_count}")
        self.stdout.write(f"  Cliente: {NOME_CLIENTE} (ID={cliente.id})")
        self.stdout.write("=" * 60)
