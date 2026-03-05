"""
Management command para importar planilhas LPU (Lista de Preços Unitários)
para o modelo Produto.

Uso:
    python manage.py importar_lpu                    # importa tudo de data/LPU/
    python manage.py importar_lpu --arquivo "LPU ADF.xlsx"  # importa apenas uma planilha
    python manage.py importar_lpu --dry-run          # simula sem salvar
"""

import os
import unicodedata
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand

import openpyxl


# Mapeamento: nome do arquivo → categoria + tipo de processamento
LPU_CONFIG = {
    "LPU ADF.xlsx": {
        "categoria": "Autodefesa",
        "prefixo": "AUT",
        "cor": "#ef4444",
        "icone": "shield",
        "tipo": "produto",
        "header_row": 2,
        "columns": {
            "codigo": 0,     # Item
            "nome": 1,       # LPU (lista unitária de preço)
            "preco_venda": 2, # VALOR VENDA
            "preco_locacao": 3,  # VALOR LOCAÇÃO 2025
        },
    },
    "LPU CFTV E ALARME.xlsx": {
        "categoria": "CFTV e Alarme",
        "prefixo": "CFT",
        "cor": "#3b82f6",
        "icone": "videocam",
        "tipo": "produto",
        "header_row": 2,
        "columns": {
            "codigo": 0,
            "nome": 1,
            "preco_venda": 2,
            "preco_locacao": 3,
        },
    },
    "LPU CONTROLE DE ACESSO.xlsx": {
        "categoria": "Controle de Acesso",
        "prefixo": "CON",
        "cor": "#8b5cf6",
        "icone": "fingerprint",
        "tipo": "produto",
        "header_row": 2,
        "columns": {
            "nome": 0,       # Descrição
            "marca": 1,      # Marca
            "modelo": 2,     # Modelo
            "preco_venda": 3, # Venda (material)
            "preco_locacao": 4,
        },
    },
    "LPU FRETE.xlsx": {
        "categoria": "Frete",
        "prefixo": "FRE",
        "cor": "#f59e0b",
        "icone": "local_shipping",
        "tipo": "servico",
        "header_row": 3,
        "columns": {
            "codigo": 0,
            "nome": 1,      # Regiões
            "preco_venda": 2, # Valor Unit. 2025
        },
    },
    "LPU OUTROS SERVIÇOS.xlsx": {
        "categoria": "Outros Serviços",
        "prefixo": "OUT",
        "cor": "#6b7280",
        "icone": "build",
        "tipo": "servico",
        "header_row": 2,
        "columns": {
            "codigo": 0,
            "nome": 1,
            "preco_venda": 2,
        },
    },
    "LPU PORTA DE ENROLAR.xlsx": {
        "categoria": "Porta de Enrolar",
        "prefixo": "POR",
        "cor": "#14b8a6",
        "icone": "door_sliding",
        "tipo": "produto",
        "header_row": 2,
        "columns": {
            "nome": 0,
            "marca": 1,
            "modelo": 2,
            "preco_venda": 3,
            "preco_locacao": 4,
        },
    },
    "LPU SERVIÇOS INFRA.xlsx": {
        "categoria": "Serviços Infra",
        "prefixo": "INF",
        "cor": "#0ea5e9",
        "icone": "construction",
        "tipo": "servico",
        "header_row": 2,
        "columns": {
            "codigo": 0,
            "nome": 1,      # Modelo (usado como nome)
            "descricao": 2,  # Descrição
            "preco_venda": 3,
        },
    },
    "LPU SERVIÇOS REAJUSTE.xlsx": {
        "categoria": "Serviços Reajuste",
        "prefixo": "REA",
        "cor": "#ec4899",
        "icone": "trending_up",
        "tipo": "servico",
        "header_row": 4,
        "columns": {
            "nome": 0,           # DESCRIÇÃO
            "preco_venda": 2,    # VALOR MENSAL POR AGÊNCIA 2025
        },
    },
    "LPU SERVIÇOS.xlsx": {
        "categoria": "Serviços Segurança",
        "prefixo": "SEG",
        "cor": "#22c55e",
        "icone": "engineering",
        "tipo": "servico",
        "header_row": 3,
        "columns": {
            "codigo": 0,     # Item
            "nome": 1,       # Modelo
            "descricao": 2,  # Descrição
            "preco_venda": 3, # RMSP 2025
        },
    },
}


def safe_decimal(value):
    """Converte valor para Decimal de forma segura."""
    if value is None:
        return Decimal("0")
    try:
        if isinstance(value, str):
            value = value.replace(",", ".").strip()
            if value.startswith("="):
                return Decimal("0")
        d = Decimal(str(value))
        return d if d >= 0 else Decimal("0")
    except (InvalidOperation, ValueError):
        return Decimal("0")


def safe_str(value, max_length=200):
    """Converte valor para string segura."""
    if value is None:
        return ""
    s = str(value).strip()
    return s[:max_length]


class Command(BaseCommand):
    help = "Importa planilhas LPU (Lista de Preços Unitários) para o modelo Produto"

    def add_arguments(self, parser):
        parser.add_argument(
            "--arquivo",
            type=str,
            help="Nome do arquivo específico para importar (ex: 'LPU ADF.xlsx')",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Simular importação sem salvar no banco",
        )
        parser.add_argument(
            "--diretorio",
            type=str,
            default="data/LPU",
            help="Diretório com as planilhas LPU (padrão: data/LPU)",
        )

    def handle(self, *args, **options):
        from dashboard.models import CategoriaEstoque, Produto, UnidadeMedida

        dry_run = options["dry_run"]
        lpu_dir = options["diretorio"]
        arquivo_filtro = options.get("arquivo")

        if not os.path.isdir(lpu_dir):
            self.stderr.write(self.style.ERROR(f"Diretório não encontrado: {lpu_dir}"))
            return

        # Garantir UnidadeMedida padrão
        un, _ = UnidadeMedida.objects.get_or_create(
            sigla="UN",
            defaults={"nome": "Unidade", "tipo": "unidade"},
        )
        sv, _ = UnidadeMedida.objects.get_or_create(
            sigla="SV",
            defaults={"nome": "Serviço", "tipo": "outros"},
        )

        total_criados = 0
        total_atualizados = 0
        total_ignorados = 0

        for fname, config in LPU_CONFIG.items():
            if arquivo_filtro and fname != arquivo_filtro:
                continue

            # Tentar encontrar o arquivo com normalização Unicode (macOS NFD)
            filepath = os.path.join(lpu_dir, fname)
            if not os.path.isfile(filepath):
                # Buscar por nome normalizado no diretório
                found = False
                for real_name in os.listdir(lpu_dir):
                    if unicodedata.normalize("NFC", real_name) == unicodedata.normalize("NFC", fname):
                        filepath = os.path.join(lpu_dir, real_name)
                        found = True
                        break
                if not found:
                    self.stdout.write(self.style.WARNING(f"Arquivo não encontrado: {fname}"))
                    continue

            self.stdout.write(f"\n{'='*60}")
            self.stdout.write(self.style.HTTP_INFO(f"Processando: {fname}"))

            # Criar ou obter categoria
            cat, created = CategoriaEstoque.objects.get_or_create(
                nome=config["categoria"],
                defaults={
                    "cor": config["cor"],
                    "icone": config["icone"],
                    "controla_estoque": config["tipo"] == "produto",
                },
            )
            if created and not dry_run:
                self.stdout.write(f"  Categoria criada: {cat.nome}")

            unidade = un if config["tipo"] == "produto" else sv

            # Ler planilha
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            header_row = config["header_row"]
            cols = config["columns"]
            criados = 0
            atualizados = 0
            ignorados = 0

            # Prefixo para código único baseado na config
            prefixo = config["prefixo"]

            for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=1):
                # Verificar se a row tem dados no campo nome
                nome_col = cols.get("nome", 0)
                if nome_col >= len(row) or not row[nome_col]:
                    ignorados += 1
                    continue

                nome = safe_str(row[nome_col])
                if not nome or nome.startswith("LPU") or nome.startswith("Item") or nome in ("Descrição", "DESCRIÇÃO"):
                    ignorados += 1
                    continue

                # Extrair código
                codigo_col = cols.get("codigo")
                if codigo_col is not None and codigo_col < len(row) and row[codigo_col]:
                    codigo_raw = safe_str(row[codigo_col], 30)
                    if codigo_raw in ("Item", "ITEM"):
                        ignorados += 1
                        continue
                    codigo = f"LPU-{prefixo}-{codigo_raw}"
                else:
                    codigo = f"LPU-{prefixo}-{row_idx:04d}"

                # Garantir código único (máx 50 chars)
                codigo = codigo[:50]

                # Extrair preço de venda
                preco_col = cols.get("preco_venda")
                preco_venda = Decimal("0")
                if preco_col is not None and preco_col < len(row):
                    preco_venda = safe_decimal(row[preco_col])

                # Extrair preço de locação
                preco_loc_col = cols.get("preco_locacao")
                preco_locacao = Decimal("0")
                if preco_loc_col is not None and preco_loc_col < len(row):
                    preco_locacao = safe_decimal(row[preco_loc_col])

                # Se não tem preço e não tem nome relevante, ignorar
                if preco_venda == 0 and preco_locacao == 0 and len(nome) < 3:
                    ignorados += 1
                    continue

                # Construir descrição
                desc_parts = []
                desc_col = cols.get("descricao")
                if desc_col is not None and desc_col < len(row) and row[desc_col]:
                    desc_parts.append(safe_str(row[desc_col], 500))
                marca_col = cols.get("marca")
                if marca_col is not None and marca_col < len(row) and row[marca_col]:
                    desc_parts.append(f"Marca: {safe_str(row[marca_col], 100)}")
                modelo_col = cols.get("modelo")
                if modelo_col is not None and modelo_col < len(row) and row[modelo_col]:
                    desc_parts.append(f"Modelo: {safe_str(row[modelo_col], 100)}")
                descricao = " | ".join(desc_parts)

                if dry_run:
                    loc_info = f" | Loc: R$ {preco_locacao:.2f}" if preco_locacao > 0 else ""
                    self.stdout.write(
                        f"  [DRY-RUN] {codigo} | {nome[:50]} | Venda: R$ {preco_venda:.2f}{loc_info}"
                    )
                    criados += 1
                    continue

                # Criar ou atualizar produto
                produto, created = Produto.objects.update_or_create(
                    codigo=codigo,
                    defaults={
                        "nome": nome[:200],
                        "descricao": descricao,
                        "tipo": config["tipo"],
                        "categoria": cat,
                        "unidade_medida": unidade,
                        "preco_venda": preco_venda,
                        "preco_locacao": preco_locacao,
                        "preco_custo": Decimal("0"),
                        "controla_estoque": config["tipo"] == "produto",
                        "status": "ativo",
                    },
                )
                if created:
                    criados += 1
                else:
                    atualizados += 1

            wb.close()

            self.stdout.write(f"  ✓ Criados: {criados} | Atualizados: {atualizados} | Ignorados: {ignorados}")
            total_criados += criados
            total_atualizados += atualizados
            total_ignorados += ignorados

        self.stdout.write(f"\n{'='*60}")
        prefix = "[DRY-RUN] " if dry_run else ""
        self.stdout.write(
            self.style.SUCCESS(
                f"{prefix}Total: {total_criados} criados, {total_atualizados} atualizados, {total_ignorados} ignorados"
            )
        )
