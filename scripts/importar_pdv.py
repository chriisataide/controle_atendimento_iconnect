#!/usr/bin/env python
"""
Script de importacao massiva de Pontos de Venda a partir da planilha
'Pontos de vendas.xlsx'.

Importa TODOS os registros (qualquer status).
Para linhas sem CNPJ, gera CNPJ ficticio baseado no UNIORG.
Para CNPJs duplicados, adiciona sufixo incremental.
"""

import os
import re
import sys

import django

# Setup Django
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, BASE_DIR)
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "controle_atendimento.settings")
django.setup()

import openpyxl
from django.db import transaction

from dashboard.models import Cliente, PontoDeVenda

# --- Configuracao ---
PLANILHA = os.path.join(BASE_DIR, "Pontos de vendas.xlsx")
NOME_CLIENTE = "Santander"
EMAIL_CLIENTE = "contato@santander.com.br"

# Placeholders para campos obrigatorios ausentes na planilha
PLACEHOLDER_EMAIL = "pdv@santander.com.br"
PLACEHOLDER_CELULAR = "(00) 00000-0000"
PLACEHOLDER_RESPONSAVEL_NOME = "A definir"
PLACEHOLDER_RESPONSAVEL_CPF = "000.000.000-00"
PLACEHOLDER_RESPONSAVEL_CARGO = "A definir"
PLACEHOLDER_RESPONSAVEL_TELEFONE = "(00) 00000-0000"
PLACEHOLDER_RESPONSAVEL_EMAIL = "responsavel@santander.com.br"


def formatar_cnpj(cnpj_raw):
    """Formata CNPJ para o padrao XX.XXX.XXX/XXXX-XX."""
    if not cnpj_raw:
        return ""
    digits = re.sub(r"\D", "", str(cnpj_raw))
    if not digits:
        return ""
    digits = digits.zfill(14)
    digits = digits[-14:]
    return "{}.{}.{}/{}-{}".format(digits[:2], digits[2:5], digits[5:8], digits[8:12], digits[12:14])


def gerar_cnpj_ficticio(uniorg, idx):
    """Gera um CNPJ ficticio baseado no UNIORG."""
    # Remove tracos do UNIORG: '001-0001' -> '0010001'
    digits = re.sub(r"\D", "", str(uniorg))
    if not digits:
        digits = str(idx)
    # Pad to 8 digits, use idx as filial
    digits = digits.zfill(8)[:8]
    filial = str(idx).zfill(4)[:4]
    check = "99"  # Digitos verificadores ficticios
    cnpj = "{}.{}.{}/{}-{}".format(digits[:2], digits[2:5], digits[5:8], filial, check)
    return cnpj


def formatar_cep(cep_raw):
    """Formata CEP para XXXXX-XXX."""
    if not cep_raw:
        return "00000-000"
    digits = re.sub(r"\D", "", str(cep_raw))
    digits = digits.zfill(8)
    return "{}-{}".format(digits[:5], digits[5:8])


def separar_logradouro_numero(endereco):
    """Separa numero do logradouro."""
    if not endereco:
        return ("S/N", "S/N")
    endereco = str(endereco).strip()
    match = re.match(r"^(.+?)\s+(\d+[\w\-]*)$", endereco)
    if match:
        return (match.group(1).strip(), match.group(2).strip())
    return (endereco, "S/N")


def formatar_telefone(fone, ddd):
    """Formata telefone com DDD."""
    if not fone:
        return ""
    fone_str = str(fone).strip()
    ddd_str = str(int(ddd)) if ddd else ""
    if ddd_str:
        return "({}) {}".format(ddd_str, fone_str)
    return fone_str


def main():
    print("=" * 60)
    print("  IMPORTACAO MASSIVA DE PONTOS DE VENDA (COMPLETA)")
    print("=" * 60)

    # 1. Abrir planilha
    print("\nAbrindo planilha: {}".format(PLANILHA))
    wb = openpyxl.load_workbook(PLANILHA, read_only=True)
    ws = wb.active
    total_linhas = ws.max_row - 1
    print("   Total de linhas: {}".format(total_linhas))

    # 2. Criar ou buscar Cliente
    cliente, created = Cliente.objects.get_or_create(
        nome=NOME_CLIENTE,
        defaults={
            "email": EMAIL_CLIENTE,
            "segmento": "Financeiro",
            "empresa": "Santander Brasil",
            "ativo": True,
        },
    )
    if created:
        print('\nCliente "{}" criado (ID={})'.format(NOME_CLIENTE, cliente.id))
    else:
        print('\nCliente "{}" ja existe (ID={})'.format(NOME_CLIENTE, cliente.id))

    # 3. Limpar PdVs existentes deste cliente para reimportar tudo
    existing_count = PontoDeVenda.objects.filter(cliente=cliente).count()
    if existing_count > 0:
        print("   Removendo {} PdVs existentes do cliente {}...".format(existing_count, NOME_CLIENTE))
        PontoDeVenda.objects.filter(cliente=cliente).delete()
        print("   Removidos com sucesso.")

    # 4. Ler e preparar dados
    cnpjs_usados = set(PontoDeVenda.objects.values_list("cnpj", flat=True))
    pdvs_para_criar = []
    cnpj_gerado_count = 0
    cnpj_dedup_count = 0
    erros = []

    print("\nProcessando linhas...")
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        uniorg = str(row[0] or "").strip()
        tipo = str(row[1] or "").strip()
        nome = str(row[2] or "").strip()
        endereco = str(row[3] or "").strip()
        cep_raw = str(row[4] or "").strip()
        bairro = str(row[5] or "").strip()
        municipio = str(row[6] or "").strip()
        uf = str(row[7] or "").strip()
        cnpj_raw = str(row[8] or "").strip()
        fone = row[9]
        ddd = row[10]
        status = str(row[11] or "").strip()
        str(row[12] or "").strip()
        rede = str(row[13] or "").strip()
        str(row[14] or "").strip()
        str(row[15] or "").strip()

        # Formatar CNPJ ou gerar ficticio
        cnpj = formatar_cnpj(cnpj_raw)
        if not cnpj or cnpj == "00.000.000/0000-00":
            cnpj = gerar_cnpj_ficticio(uniorg, idx)
            cnpj_gerado_count += 1

        # Resolver duplicatas adicionando sufixo
        original_cnpj = cnpj
        suffix = 1
        while cnpj in cnpjs_usados:
            # Modifica os ultimos 2 digitos (verificadores)
            base = original_cnpj[:-2]
            cnpj = "{}{:02d}".format(base, suffix)
            suffix += 1
            cnpj_dedup_count += 1

        cnpjs_usados.add(cnpj)

        # Separar logradouro e numero
        logradouro, numero = separar_logradouro_numero(endereco)

        # Montar nome fantasia
        nome_fantasia = nome if nome else "{} {}".format(tipo, uniorg).strip()
        razao_social = "{} ({} {})".format(nome, tipo, uniorg).strip() if nome else nome_fantasia

        # Telefone formatado
        telefone = formatar_telefone(fone, ddd)

        # Complemento com tipo, rede, reg e status
        complemento_parts = [tipo]
        if rede:
            complemento_parts.append(rede)
        comp = " | ".join(complemento_parts)

        try:
            pdv = PontoDeVenda(
                cliente=cliente,
                razao_social=razao_social[:150],
                nome_fantasia=nome_fantasia[:150],
                cnpj=cnpj,
                inscricao_estadual=uniorg[:30],  # Guardar UNIORG aqui
                inscricao_municipal=status[:30],  # Guardar Status original aqui
                cep=formatar_cep(cep_raw),
                logradouro=logradouro[:120],
                numero=numero[:10],
                complemento=comp[:50],
                bairro=bairro[:60] if bairro else "N/I",
                cidade=municipio[:60] if municipio else "N/I",
                estado=uf[:2] if uf else "XX",
                pais="Brasil",
                celular=telefone if telefone else PLACEHOLDER_CELULAR,
                email_principal=PLACEHOLDER_EMAIL,
                email_financeiro="",
                website="",
                responsavel_nome=PLACEHOLDER_RESPONSAVEL_NOME,
                responsavel_cpf=PLACEHOLDER_RESPONSAVEL_CPF,
                responsavel_cargo=PLACEHOLDER_RESPONSAVEL_CARGO,
                responsavel_telefone=PLACEHOLDER_RESPONSAVEL_TELEFONE,
                responsavel_email=PLACEHOLDER_RESPONSAVEL_EMAIL,
            )
            pdvs_para_criar.append(pdv)
        except Exception as e:
            erros.append("Linha {}: {}".format(idx, e))

    print("   Pronto para importar: {}".format(len(pdvs_para_criar)))
    print("   CNPJs gerados (ficticios): {}".format(cnpj_gerado_count))
    print("   CNPJs deduplicados: {}".format(cnpj_dedup_count))
    if erros:
        print("   Erros: {}".format(len(erros)))
        for e in erros[:10]:
            print("     ERRO: {}".format(e))

    if not pdvs_para_criar:
        print("\nNenhum ponto de venda para importar.")
        return

    # 5. Criptografar campos PII antes do bulk_create
    print("\nCriptografando campos PII...")
    try:
        from dashboard.utils.crypto import encrypt_value

        for pdv in pdvs_para_criar:
            if pdv.responsavel_cpf and not pdv.responsavel_cpf.startswith("enc::"):
                pdv.responsavel_cpf = encrypt_value(pdv.responsavel_cpf)
            if pdv.celular and not pdv.celular.startswith("enc::"):
                pdv.celular = encrypt_value(pdv.celular)
            if pdv.responsavel_telefone and not pdv.responsavel_telefone.startswith("enc::"):
                pdv.responsavel_telefone = encrypt_value(pdv.responsavel_telefone)
    except ImportError:
        print("   Modulo crypto nao disponivel, salvando sem criptografia")

    # 6. Inserir em lote
    print("\nInserindo {} pontos de venda...".format(len(pdvs_para_criar)))
    BATCH_SIZE = 500
    total_criados = 0

    with transaction.atomic():
        for i in range(0, len(pdvs_para_criar), BATCH_SIZE):
            batch = pdvs_para_criar[i : i + BATCH_SIZE]
            PontoDeVenda.objects.bulk_create(batch, ignore_conflicts=True)
            total_criados += len(batch)
            print("   ... {}/{} inseridos".format(total_criados, len(pdvs_para_criar)))

    wb.close()

    # 7. Resumo
    total_final = PontoDeVenda.objects.count()
    print("\n" + "=" * 60)
    print("  RESUMO DA IMPORTACAO")
    print("=" * 60)
    print("  Pontos de venda criados: {}".format(len(pdvs_para_criar)))
    print("  Total de PdVs no banco: {}".format(total_final))
    print("  CNPJs ficticios gerados: {}".format(cnpj_gerado_count))
    print("  Cliente: {} (ID={})".format(NOME_CLIENTE, cliente.id))
    print("=" * 60)


if __name__ == "__main__":
    main()
